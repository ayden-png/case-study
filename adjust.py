import openpyxl

filename = "C:\\Users\shang\OneDrive\Desktop\\100 US Mid-Cap Stocks.xlsx"

workbook = openpyxl.load_workbook(filename)
worksheet1 = workbook.active

insertrows = []
rows = worksheet1.max_row

print(rows)
print(worksheet1.max_column)

mean = 0
count = 0
summary = 0
list_mean = []
vol_sum = 0
count1 =0
list_vol = []
vol = 0
a = 0

for i in range(rows-1,0,-1):
    if worksheet1.cell(row=i+1,column=2).value != worksheet1.cell(row=i,column=2).value:
        insertrows.append(i+1)

print(insertrows)

for i in range(0,len(insertrows)-1,1):

    worksheet1.insert_rows(insertrows[i])

    if i != len(insertrows)-2:
        for j in range(insertrows[i] - 1, insertrows[i + 1]-1, -1):
            if type(worksheet1.cell(row=j, column=6).value) == float or type(
                    worksheet1.cell(row=j, column=6).value) == int :
                summary += worksheet1.cell(row=j, column=6).value
                count += 1
        mean = summary / count
        worksheet1.cell(row=insertrows[i], column=6).value = mean
        list_mean.append(mean)
        summary = 0
        count = 0

        for k in range(insertrows[i] - 1, insertrows[i + 1]-1, -1):
            if type(worksheet1.cell(row=k, column=6).value) == float or type(
                    worksheet1.cell(row=k, column=6).value) == int:
                a = (worksheet1.cell(row=k,column=6).value-mean)**2
                vol_sum += a
                count1 += 1
        vol = vol_sum/(count1-1)
        vol = vol**(1/2)
        worksheet1.cell(row=insertrows[i], column=7).value = vol
        list_vol.append(vol)
        vol_sum = 0
        count1 = 0

workbook.create_sheet('market return',1)
worksheet2 = workbook.worksheets[1]
list_mean.append('none')
list_vol.append('none')
worksheet2.cell(row=1,column=1).value = 'Date No.'
worksheet2.cell(row=1,column=2).value = 'daily average return'

for i in range(0,253,1):
    worksheet2.cell(row=i + 2, column=1).value = 253-i
    worksheet2.cell(row=i + 2, column=2).value = list_mean[i]
    worksheet2.cell(row=i + 2, column=3).value = list_vol[i]

workbook.save(filename=filename)